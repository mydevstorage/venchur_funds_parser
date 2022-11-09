import re
import time
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
import json
import csv
from openpyxl import load_workbook, Workbook
import random
import os
import sqlite3
from selenium.webdriver.support.wait import WebDriverWait
# from loguru import logger
import argparse

NO_INFO_STATUS = 'No information'

DATA_FOLDER = '/home/roman/real_python/web_parsing/venchur_funds_parser'

PATH_TO_CHROME_DRIVER = ("/home/roman/real_python/web_parsing/"
                         "venchur_funds_parser/chrome_ driver/chromedriver")

user_agent = ("user-agent=Mozilla/5.0 (X11; Linux x86_64) '\
                         'AppleWebKit/537.36 (KHTML, like Gecko) '\
                         'Chrome/106.0.0.0 Safari/537.36")

PATH_TO_FIREFOX_DRIVER = ("/home/roman/real_python/web_parsing/"
                          "venchur_funds_parser/firefox_driver/geckodriver")

# Для анализа всех данных = 14840, займет часов 11 для анализа
AMOUNT_OF_FUNDS_FOR_PARSING = 10

AMOUNT_MANAGGERS_IN_EXCEL_CSV_TABLE = 10

STEP = 10       # Шаг, на каждой странице по 10 фондов/ссылок

HEADERS_RESULT_TABLE = (
    'Name of investor',
    'Site',
    'Info card',
    'Stage',
    'Check size',
    'Focus',
    'Investment geography',
    'Manager name', 'Role', 'Contact',
    'Manager name', 'Role', 'Contact',
    'Manager name', 'Role', 'Contact',
    'Manager name', 'Role', 'Contact',
    'Manager name', 'Role', 'Contact',
    'Manager name', 'Role', 'Contact'
    'Manager name', 'Role', 'Contact',
    'Manager name', 'Role', 'Contact',
    'Manager name', 'Role', 'Contact',
    'Manager name', 'Role', 'Contact')


def get_options():

    parser = argparse.ArgumentParser()
    HELP_INFO1 = 'choose and type [csv, excel ,db] for writing method'
    HELP_INFO2 = 'choose and type [chrome, firefox] for browser tool'

    parser.add_argument('output_format', choices=['csv', 'excel', 'sqlite3'],
                        help=HELP_INFO1)
    parser.add_argument('browser', choices=['chrome',
                        'firefox'], help=HELP_INFO2)
    return parser.parse_args()


def get_all_links():

    headers = {"accept": "text/css,*/*;q=0.1",
               'user-agent': 'Mozilla/5.0 (X11; Linux x86_64) '
               'AppleWebKit/537.36 (KHTML, like Gecko) '
               'Chrome/106.0.0.0 Safari/537.36'}

    for page in range(0, AMOUNT_OF_FUNDS_FOR_PARSING, STEP):
        fund_links = {}
        gfs = requests.get(f"https://project-valentine-api.herokuapp.com/"
                           f"investors?page%5Blimit%5D="
                           f"10&page%5Boffset%5D={page}", headers=headers)
        s = gfs.json()
        for i in s['data']:
            temp = (f"https://connect.visible.vc/investors/"
                    f"{i['attributes']['slug']}")
            name = i['attributes']['name']
            fund_links[name] = temp

        if not os.path.exists(f"{DATA_FOLDER}/data"):
            os.mkdir(f"{DATA_FOLDER}/data")

        with open(f"{DATA_FOLDER}/data/all_links_{page}.json", "w") as file:
            json.dump(fund_links, file, indent=4, ensure_ascii=False)

        time.sleep(random.randint(1, 2))

        print(f'Progress ... {page}')


def create_headers_in_csv_table():

    with open(f"{DATA_FOLDER}/result.csv", 'w', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerow(HEADERS_RESULT_TABLE)


def create_headers_in_excel_table():

    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS_RESULT_TABLE)
    wb.save(f"{DATA_FOLDER}/result.xlsx")
    wb.close()


def create_database_tables():

    con = sqlite3.connect(f"{DATA_FOLDER}/result.db")
    cur = con.cursor()
    cur.execute("PRAGMA busy_timeout = 30000")
    cur.execute('''CREATE TABLE if not exists INVESTORS(
            investor_id integer,
            name_of_investor text,
            site text,
            info_card text,
            stage text text,
            check_size text,
            focus text,
            investment_geography text
    )''')

    cur.execute('''CREATE TABLE if not exists MANAGERS(
            manager_id integer primary key autoincrement,
            investor_id integer,
            manager_name text,
            role text,
            contacts text,
            foreign key (investor_id) references
            INVESTORS(investor_id) on delete cascade
    )''')
    con.commit()
    con.close()


def get_driver_chrome():

    options = webdriver.ChromeOptions()
    options.add_argument(user_agent)
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.headless = True
    driver = webdriver.Chrome(executable_path=PATH_TO_CHROME_DRIVER,
                              options=options)
    return driver
# firefox_profile='/home/roman/snap/firefox/common/.mozilla/firefox/profiles.ini'


def get_driver_firefox():

    options = webdriver.FirefoxOptions()
    options.set_preference("general.useragent.override", user_agent)
    options.set_preference("dom.webdriver.enabled", False)
    options.headless = True
    driver = webdriver.Firefox(executable_path=PATH_TO_FIREFOX_DRIVER,
                               options=options)
    return driver


def treatment_of_data_with_browser(outer_args):
    count_of_funds = 1
    try:
        if outer_args.browser == 'chrome':
            driver = get_driver_chrome()
        elif outer_args.browser == 'firefox':
            driver = get_driver_firefox()
        for page in range(0, AMOUNT_OF_FUNDS_FOR_PARSING, STEP):
            with open(f"{DATA_FOLDER}/data/all_links_{page}.json") as file:
                all_links = json.load(file)
            get_data_from_pages(all_links, driver, count_of_funds, outer_args)
    except Exception as ex:
        print(ex)
    finally:
        driver.close()
        driver.quit()


def get_data_from_pages(all_links, driver, count_of_funds, outer_args):

    for name_fund in all_links:

        try:
            driver.get(url=all_links[name_fund])
            time.sleep(2)
            WebDriverWait(driver, 20).until(
                          lambda x: x.find_element(By.ID, "ember7"))

            soup = BeautifulSoup(driver.page_source, 'lxml')
            print(f'{all_links[name_fund]} ... Готов!')

            link = soup.find(class_='mr-2 text-sm leading-tight '
                                    'text-orange-600 hover:'
                                    'underline').get('href').strip(' ')
            stage = (soup.find(string=re.compile('Stage'))
                         .find_next('span').text.replace('\n', '')
                         .replace(' ', ''))
            check = (soup.find(string=re.compile('Check size'))
                         .find_next('span')
                         .text.replace('\n', '')
                         .replace(' ', ''))
            focus = (soup.find(string=re.compile('Focus'))
                         .find_next('span').string.replace(' ', '')
                         .replace(',', ', ').strip('\n '))
            i_geo = (soup.find(string=re.compile('Investment geography'))
                         .find_next('span')
                         .text.strip().replace('\n', ''))

            all_names = []
            all_roles = []
            all_contacts = []

            m_name = soup.find_all(class_='font-serif '
                                   'text-base text-black truncate')
            for i in m_name:
                all_names.append(i.text.strip('\n '))

            m_role = soup.find_all(class_='text-xs text-gray-500 truncate')
            for i in m_role:
                all_roles.append(i.text.strip('\n ')
                                 if i.text.strip('\n ') != ''
                                 else NO_INFO_STATUS)

            m_links = soup.find_all(class_='flex items-center px-2 py-3 '
                                    'text-sm bg-white border '
                                    'border-gray-300 max-w-sm')
            for i in m_links:
                tmp = ''
                for e in i.find_all('a'):
                    tmp += e.get('href') + ', '
                all_contacts.append(tmp.strip('\n')
                                    if tmp.strip('\n') != ''
                                    else NO_INFO_STATUS)

            main_row_for_table = [
                    count_of_funds,
                    name_fund,
                    (link if link != '' else NO_INFO_STATUS),
                    all_links[name_fund],
                    (stage if stage != '' else NO_INFO_STATUS),
                    (check if check != '-' else NO_INFO_STATUS),
                    (focus if focus != '' else NO_INFO_STATUS),
                    (i_geo if i_geo != '' else NO_INFO_STATUS)]

            for i in range(AMOUNT_MANAGGERS_IN_EXCEL_CSV_TABLE):
                try:
                    main_row_for_table.append(all_names[i])
                    main_row_for_table.append(all_roles[i])
                    main_row_for_table.append(all_contacts[i])
                except Exception:
                    main_row_for_table.append(NO_INFO_STATUS)
                    main_row_for_table.append(NO_INFO_STATUS)
                    main_row_for_table.append(NO_INFO_STATUS)

            if outer_args.output_format == 'csv':
                append_data_to_csv(main_row_for_table)
            elif outer_args.output_format == 'excel':
                append_data_to_excel(main_row_for_table)
            elif outer_args.output_format == 'sqlite3':
                append_data_to_database(all_names, all_roles, all_contacts,
                                        main_row_for_table, count_of_funds)

            count_of_funds += 1   # Прибавляем счетчик инвесторов (investor_id)
        except Exception as ex:
            print(ex)


def append_data_to_database(all_names, all_roles, all_contacts,
                            main_row_for_table, count_of_funds):
    con = sqlite3.connect(f"{DATA_FOLDER}/result.db")
    cur = con.cursor()
    cur.executemany('''insert into INVESTORS(
                        investor_id,
                        name_of_investor,
                        site,
                        info_card,
                        stage,
                        check_size,
                        focus,
                        investment_geography)
                        values(?, ?, ?, ?, ?, ?, ?, ?)''',
                    (main_row_for_table[:8],))
    con.commit()

    for num in range(len(all_names)):
        row_for_database_table_MANAGERS = [count_of_funds, all_names[num],
                                           all_roles[num], all_contacts[num]]
        cur.executemany('''insert into MANAGERS(
                            investor_id,
                            manager_name,
                            role,
                            contacts) values(?, ?, ?, ?)''',
                        (row_for_database_table_MANAGERS,))

    con.commit()
    con.close()


def append_data_to_csv(main_row_for_table):
    del main_row_for_table[0]
    with open(f"{DATA_FOLDER}/result.csv", 'a') as file:
        writer = csv.writer(file)
        writer.writerow(main_row_for_table)


def append_data_to_excel(main_row_for_table):
    del main_row_for_table[0]
    wb = load_workbook(f"{DATA_FOLDER}/result.xlsx")
    ws = wb.active
    ws.append(main_row_for_table)
    wb.save(f"{DATA_FOLDER}/result.xlsx")


def main():
    outer_args = get_options()

    if outer_args.output_format == 'csv':
        create_headers_in_csv_table()
    elif outer_args.output_format == 'excel':
        create_headers_in_excel_table()
    elif outer_args.output_format == 'sqlite3':
        create_database_tables()

    get_all_links()
    treatment_of_data_with_browser(outer_args)


if __name__ == '__main__':
    main()
