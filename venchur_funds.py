import re
import time
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
import json
import csv
from openpyxl import load_workbook, Workbook
import random
import  os
import sqlite3
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support.expected_conditions import visibility_of_element_located
from datetime import datetime

NO_INFO_STATUS = 'No information'

DATA_FOLDER = 'Parsing/Venchur_fonds/'

PATH_TO_WEBDRIVER = "/home/roman/python_course/Selenium_Python/Chrome_ driver/chromedriver"

AMOUNT_OF_FUNDS_FOR_PARSING = 10   # Для анализа всех данных = 14840, займет часов 11 для анализа

AMOUNT_MANAGGERS_IN_EXCEL_CSV_TABLE = 10

STEP = 10       # Шаг, на каждой странице по 10 фондов/ссылок

COUNT_OF_FUNDS = 1



def get_all_links():
    ''' Функция собирает ссылки на каждый фонд и записывает их в json файл. '''

    headers = {"accept": "text/css,*/*;q=0.1", 'user-agent' : 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/106.0.0.0 Safari/537.36'}

    for page in range(0, AMOUNT_OF_FUNDS_FOR_PARSING, STEP): 
        fund_links = {}   
        gfs = requests.get(f'https://project-valentine-api.herokuapp.com/investors?page%5Blimit%5D=10&page%5Boffset%5D={page}',
        headers=headers)

        s = gfs.json()    

        for i in s['data']:
            temp = f'https://connect.visible.vc/investors/{i["attributes"]["slug"]}'
            name = i['attributes']['name']
            fund_links[name] = temp

        if not os.path.exists(f"{DATA_FOLDER}/data"):
            os.mkdir(f"{DATA_FOLDER}data")

        with open(f"{DATA_FOLDER}/data/all_links_{page}.json", "w") as file:
            json.dump(fund_links, file, indent=4, ensure_ascii=False)

        time.sleep(random.randint(1, 2))

        print(f'Progress ... {page}')

def create_headers_in_csv_table():
    ''' Создание файла CSV и добавление в него заголовков. База вмещает информацию о 10 сотрудниках максимум'''

    global HEADERS_RESULT_TABLE

    HEADERS_RESULT_TABLE = (
        'Name of investor',
        'Site',
        'Info card',
        'Stage',
        'Check size',
        'Focus',
        'Investment geography',
        'Manager name',
        'Role',
        'Contact',
        'Manager name', 'Role', 'Contact', 'Manager name', 'Role', 'Contact', 'Manager name', 'Role', 'Contact',
        'Manager name', 'Role', 'Contact', 'Manager name', 'Role', 'Contact', 'Manager name', 'Role', 'Contact',
        'Manager name', 'Role', 'Contact', 'Manager name', 'Role', 'Contact', 'Manager name', 'Role', 'Contact')

    with open(f"{DATA_FOLDER}/result.csv", 'w', encoding='utf-8') as file:  # Добавление заголовков в таблицу csv
        writer = csv.writer(file)
        writer.writerow(HEADERS_RESULT_TABLE)

def create_headers_in_excel_table():
    ''' Создание файла Excel и добавление заголовков в таблицу. '''

    wb = Workbook()                                                          
    ws = wb.active
    ws.append(HEADERS_RESULT_TABLE) 
    wb.save(f"{DATA_FOLDER}/result.xlsx")
    wb.close()

def create_database_tables():
    ''' Создание базы данных, состоящей из двух связанных таблиц. '''

    con = sqlite3.connect(f"{DATA_FOLDER}/result.db")                        
    cur = con.cursor()
    cur.execute("PRAGMA busy_timeout = 30000")                 # При множественном обращении запись идет с пропусками
    cur.execute('''CREATE TABLE if not exists INVESTORS(
            investor_id integer,
            name_of_investor text,
            site text,
            info_card text,
            stage text text,
            check_size text,
            focus text,
            investment_geography text
    )''')

    cur.execute('''CREATE TABLE if not exists MANAGERS(
            manager_id integer primary key autoincrement,
            investor_id integer,
            manager_name text,
            role text,
            contacts text,
            foreign key (investor_id) references INVESTORS(investor_id) on delete cascade
    )''')
    con.commit()
    con.close()

def treatment_of_data_with_selenium():
    
    options = webdriver.ChromeOptions()
    options.add_argument("user-agent=Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/106.0.0.0 Safari/537.36")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.headless = True
    
    try:
        global driver
        driver = webdriver.Chrome(executable_path=PATH_TO_WEBDRIVER, options=options)

        for page in range(0, AMOUNT_OF_FUNDS_FOR_PARSING, STEP): 

            with open(f"{DATA_FOLDER}/data/all_links_{page}.json") as file:
                all_links = json.load(file)     

            get_data_from_pages(all_links)

    except Exception as ex:
        print(ex)

    finally:
        driver.close()
        driver.quit()
       
def get_data_from_pages(all_links):             
    ''' Cбор данных с каждого фонда и запись их в базу данных, CSV и Excel'''

    global COUNT_OF_FUNDS
    for name_fund in all_links:
          
        try:        
            driver.get(url=all_links[name_fund])
            time.sleep(2)
            wait = WebDriverWait(driver, 20).until(lambda x: x.find_element(By.ID, "ember7"))   # Ожидание загрузки до появления конкретного id

            soup = BeautifulSoup(driver.page_source, 'lxml')         
            print(f'{all_links[name_fund]} ... Готов!')

            link = soup.find(class_='mr-2 text-sm leading-tight text-orange-600 hover:underline').get('href').strip(' ')          
            stage = soup.find(string=re.compile('Stage')).find_next('span').text.replace('\n', '').replace(' ', '')
            check = soup.find(string=re.compile('Check size')).find_next('span').text.replace('\n', '').replace(' ', '')
            focus = soup.find(string=re.compile('Focus')).find_next('span').string.replace(' ', '').replace(',', ', ').strip('\n ')
            i_geo = soup.find(string=re.compile('Investment geography')).find_next('span').text.strip().replace('\n', '')
            
            all_names =[]
            all_roles = []
            all_contacts = []

            m_name = soup.find_all(class_='font-serif text-base text-black truncate')
            for i in m_name:
                all_names.append(i.text.strip('\n '))
            
            m_role = soup.find_all(class_='text-xs text-gray-500 truncate')
            for i in m_role:
                all_roles.append(i.text.strip('\n ') if i.text.strip('\n ') != '' else NO_INFO_STATUS)
            
            m_links = soup.find_all(class_='flex items-center px-2 py-3 text-sm bg-white border border-gray-300 max-w-sm')
            for i in m_links:
                tmp = ''
                for e in i.find_all('a'):
                    tmp += e.get('href') + ', '
                all_contacts.append(tmp.strip('\n') if tmp.strip('\n') != '' else NO_INFO_STATUS)  

            main_row_for_table = [
                    COUNT_OF_FUNDS,
                    name_fund,
                    (link if link != '' else NO_INFO_STATUS),
                    all_links[name_fund],
                    (stage if stage != '' else NO_INFO_STATUS),
                    (check if check != '-' else NO_INFO_STATUS),
                    (focus if focus != '' else NO_INFO_STATUS),
                    (i_geo if i_geo != '' else NO_INFO_STATUS)]

            append_data_to_database(all_names, all_roles,all_contacts, main_row_for_table, COUNT_OF_FUNDS)

            COUNT_OF_FUNDS += 1   # Прибавляем счетчик инвесторов (investor_id)

            for i in range(AMOUNT_MANAGGERS_IN_EXCEL_CSV_TABLE):     # Добавление всех сотрудников в общий список main_row_for_table
                try:
                    main_row_for_table.append(all_names[i])
                    main_row_for_table.append(all_roles[i])
                    main_row_for_table.append(all_contacts[i])
                except: 
                    main_row_for_table.append(NO_INFO_STATUS)
                    main_row_for_table.append(NO_INFO_STATUS)
                    main_row_for_table.append(NO_INFO_STATUS)             
            
            del main_row_for_table[0]    # Удаление счетчика инвесторов для подготовки файла для excel & csv

            append_data_to_csv(main_row_for_table)
          
            append_data_to_excel(main_row_for_table)
                    
        except Exception as ex:
                print(ex)

def append_data_to_database(all_names, all_roles,all_contacts, main_row_for_table, COUNT_OF_FUNDS):
    ''' Запись данных в бд.'''
    con = sqlite3.connect(f"{DATA_FOLDER}/result.db")                    
    cur = con.cursor()
    cur.executemany('''insert into INVESTORS( 
                        investor_id,
                        name_of_investor,
                        site,
                        info_card,
                        stage,
                        check_size,
                        focus,
                        investment_geography) values(?, ?, ?, ?, ?, ?, ?, ?)''', (main_row_for_table,))
    con.commit()
   
    for num in range(len(all_names)):
        row_for_database_table_MANAGERS = [COUNT_OF_FUNDS, all_names[num], all_roles[num], all_contacts[num]]
        cur.executemany('''insert into MANAGERS(
                            investor_id,
                            manager_name,
                            role,
                            contacts 
                            ) values(?, ?, ?, ?)''', (row_for_database_table_MANAGERS,) )
    
    con.commit()
    con.close()   

def append_data_to_csv(main_row_for_table):           
    
    with open(f"{DATA_FOLDER}/result.csv", 'a') as file:        
        writer = csv.writer(file)
        writer.writerow(main_row_for_table)

def append_data_to_excel(main_row_for_table):        

    wb = load_workbook(f"{DATA_FOLDER}/result.xlsx")  
    ws= wb.active                                    
    ws.append(main_row_for_table)                    
    wb.save(f"{DATA_FOLDER}/result.xlsx")
    
def main():

    create_database_tables()
    create_headers_in_csv_table()
    create_headers_in_excel_table()
    get_all_links()
    treatment_of_data_with_selenium()

if __name__ == '__main__':
    main()